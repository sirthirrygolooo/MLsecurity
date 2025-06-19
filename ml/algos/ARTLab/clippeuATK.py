import clip
import torch
import torch.nn as nn
from PIL import Image
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from torchvision import transforms

# === Configuration ===
torch.autograd.set_detect_anomaly(False)
device = "cuda" if torch.cuda.is_available() else "cpu"
model, preprocess = clip.load("ViT-B/32", device=device)
model.eval()
print(f"[+] Utilisation de {torch.cuda.get_device_name(0)}" if device == "cuda" else "[!] Utilisation CPU")

# === Classes ===
classes = ["trafficlight", "stop", "speedlimit", "crosswalk"]
text_inputs = torch.cat([clip.tokenize(f"a photo of a {c}") for c in classes]).to(device)
text_features = model.encode_text(text_inputs).detach()
text_features = text_features / text_features.norm(dim=-1, keepdim=True)

# === Fichiers ===
image_folder = "RoadSign/images"
atk_image_folder = "atk_roadsign"
os.makedirs(atk_image_folder, exist_ok=True)

# === Excel Styles ===
success_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
failure_fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")
warning_fill = PatternFill(start_color="FFF5BA", end_color="FFF5BA", fill_type="solid")


# === Fonction DeepFool adaptée à CLIP ===
def deepfool(image, model, text_features, num_classes=4, overshoot=0.02, max_iter=50):
    image = image.clone().detach().requires_grad_(True).to(device)
    image_features = model.encode_image(image)
    image_features = image_features / image_features.norm(dim=-1, keepdim=True)
    logits = (image_features @ text_features.T).squeeze(0)
    label = logits.argmax().item()

    pert_image = image.clone()
    r_tot = torch.zeros_like(image)

    for _ in range(max_iter):
        pert_image = pert_image.clone().detach().requires_grad_(True)
        image_features = model.encode_image(pert_image)
        image_features = image_features / image_features.norm(dim=-1, keepdim=True)
        logits = (image_features @ text_features.T).squeeze(0)
        logit_label = logits[label]

        logit_label.backward(retain_graph=True)
        grad_orig = pert_image.grad.data.clone()

        min_val = float('inf')
        w = torch.zeros_like(image)

        for k in range(num_classes):
            if k == label:
                continue
            pert_image.grad.zero_()
            logits[k].backward(retain_graph=True)
            cur_grad = pert_image.grad.data.clone()

            w_k = cur_grad - grad_orig
            f_k = (logits[k] - logit_label).item()
            norm_w_k = torch.norm(w_k.flatten())
            if norm_w_k == 0:
                continue

            pert_k = abs(f_k) / norm_w_k
            if pert_k < min_val:
                min_val = pert_k
                w = w_k

        if torch.norm(w.flatten()) == 0:
            break

        r_i = (min_val + 1e-4) * w / torch.norm(w.flatten())
        r_tot += r_i
        pert_image = torch.clamp(image + (1 + overshoot) * r_tot, 0, 1).detach()

        image_features = model.encode_image(pert_image)
        image_features = image_features / image_features.norm(dim=-1, keepdim=True)
        logits = (image_features @ text_features.T).squeeze(0)
        if logits.argmax().item() != label:
            break

    return pert_image.detach(), label, logits.argmax().item()


# === Prédiction CLIP ===
def predict_class(image):
    with torch.no_grad():
        image_features = model.encode_image(image)
        image_features = image_features / image_features.norm(dim=-1, keepdim=True)
        similarity = (image_features @ text_features.T).softmax(dim=-1)
        top_val, top_idx = similarity[0].topk(1)
        return classes[top_idx.item()], top_val.item(), similarity


# === Traitement des images ===
def process_images():
    wb = Workbook()
    ws = wb.active
    ws.append(["Image", "Attacked", "Original Class", "Original Confidence",
               "Attacked Class", "Attacked Confidence", "Attack Success", "Confidence Change"])

    image_files = [f for f in os.listdir(image_folder) if f.lower().endswith(('.png', '.jpg', '.jpeg'))]
    num_images_to_attack = max(1, int(len(image_files) * 0.1))

    # === Partie Attaquée ===
    for image_file in image_files[:num_images_to_attack]:
        image_path = os.path.join(image_folder, image_file)
        image = preprocess(Image.open(image_path).convert("RGB")).unsqueeze(0).to(device)

        orig_class, orig_conf, _ = predict_class(image)
        perturbed_image, orig_idx, new_idx = deepfool(image, model, text_features)
        atk_class, atk_conf, _ = predict_class(perturbed_image)

        attack_success = (orig_class != atk_class)
        confidence_diff = atk_conf - orig_conf

        transforms.ToPILImage()(perturbed_image.squeeze(0).cpu()).save(
            os.path.join(atk_image_folder, f"atk_{image_file}")
        )

        row = [
            image_file, "Yes", orig_class, f"{orig_conf:.4f}",
            atk_class, f"{atk_conf:.4f}", "Yes" if attack_success else "No",
            f"{confidence_diff:.4f}"
        ]
        ws.append(row)

        fill = success_fill if attack_success else failure_fill
        if attack_success and confidence_diff > 0.1:
            fill = warning_fill
        for cell in ws[ws.max_row][2:]:
            cell.fill = fill

    # === Partie Non Attaquée ===
    for image_file in image_files[num_images_to_attack:num_images_to_attack * 2]:
        image_path = os.path.join(image_folder, image_file)
        image = preprocess(Image.open(image_path).convert("RGB")).unsqueeze(0).to(device)
        orig_class, orig_conf, _ = predict_class(image)
        ws.append([image_file, "No", orig_class, f"{orig_conf:.4f}", "-", "-", "-", "-"])

    # === Ajustement largeur colonnes ===
    for column in ws.columns:
        max_length = max(len(str(cell.value)) for cell in column if cell.value)
        col_letter = column[0].column_letter
        ws.column_dimensions[col_letter].width = (max_length + 2) * 1.2

    wb.save("attack_results_deepfool.xlsx")
    print("[✓] Résultats sauvegardés dans 'attack_results_deepfool.xlsx'")
    print(f"[✓] Images attaquées sauvegardées dans '{atk_image_folder}'")


if __name__ == "__main__":
    process_images()
